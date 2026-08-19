#!/usr/bin/env python3
"""
Geocoded_v4.xlsx'ten (Bina No + guvenilir location_type) sistemin
/api/v1/scenarios/import uc noktasinin bekledigi 'personel' sayfasina
(sicil numarasi, ad soyad, adres, boylam, enlem) donusturur. APPROXIMATE
location_type'lar disaridi birakilir (ayri incelenecek).

Kullanim:
    python scripts/build_import_excel.py --input "<Geocoded_v4.xlsx>" --output "<ImportReady.xlsx>"
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

RELIABLE_TYPES = {"ROOFTOP", "RANGE_INTERPOLATED", "GEOMETRIC_CENTER"}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    rows = list(rows_iter)

    ready = [r for r in rows if r[idx["location_type"]] in RELIABLE_TYPES]
    no_sicil = [r for r in ready if not r[idx["Sicil"]]]
    ready = [r for r in ready if r[idx["Sicil"]]]
    skipped = len(rows) - len(ready) - len(no_sicil)

    out_wb = openpyxl.Workbook()
    ws_out = out_wb.active
    ws_out.title = "personel"
    ws_out.append(["sicil numarası", "ad soyad", "adres", "boylam", "enlem"])
    for r in ready:
        ws_out.append([
            r[idx["Sicil"]],
            r[idx["Ad Soyad"]],
            r[idx["Adres"]],
            r[idx["lon"]],
            r[idx["lat"]],
        ])
    out_wb.save(args.output)

    print(f"Toplam: {len(rows)}")
    print(f"Yuklemeye hazir: {len(ready)}")
    print(f"Disaride birakilan (APPROXIMATE): {skipped}")
    print(f"Sicil no eksik (disaride birakildi): {len(no_sicil)}")
    for r in no_sicil:
        print(f"  {r[idx['Ad Soyad']]!r}")
    print(f"Cikti: {args.output}")


if __name__ == "__main__":
    main()
