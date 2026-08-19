#!/usr/bin/env python3
"""
"Adres neredeyse bos" (sadece Ek Konum + Ilce/Il) olan kritik kayitlar icin,
Google'in sokak adresi yerine yer/site adiyla arama yapabildigi gozlemine
dayanarak "Ek Konum, Ilce, Il" formatinda ayri bir sorgu listesi uretir. Bu
sorgular Places API - Text Search ile denenmek uzere ayri tutulur; ana
pipeline'daki (adres bazli) sorgulari etkilemez.

Kullanim:
    python scripts/build_ekkonum_only_queries.py --input "<TamAdresIste.xlsx>" --output "<EkKonumSorgu.xlsx>"
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl

EK_KONUM_PATTERN = re.compile(r"Ek Konum\s*:\s*([^,]+)", re.IGNORECASE)


def extract_ek_konum(address: str) -> str:
    match = EK_KONUM_PATTERN.search(address)
    return match.group(1).strip() if match else ""


def extract_ilce_il(address: str) -> str:
    segments = [s.strip() for s in address.split(",") if s.strip()]
    return ", ".join(segments[-2:]) if len(segments) >= 2 else ", ".join(segments)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    rows = list(rows_iter)

    out_header = [*header, "Ek Konum Sorgusu", "Sorgulanabilir mi"]
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(out_header)

    queryable = 0
    for row in rows:
        address = str(row[idx["Adres"]] or "")
        ek_konum = extract_ek_konum(address)
        ilce_il = extract_ilce_il(address)
        query = f"{ek_konum}, {ilce_il}" if ek_konum else ""
        if query:
            queryable += 1
        out_ws.append([*row, query, "Evet" if query else "Hayır"])

    out_wb.save(args.output)

    print(f"Toplam: {len(rows)}")
    print(f"Sorgulanabilir (Ek Konum var): {queryable}")
    print(f"Sorgulanamaz (hicbir bilgi yok): {len(rows) - queryable}")
    print(f"Cikti: {args.output}")


if __name__ == "__main__":
    main()
