#!/usr/bin/env python3
"""
prepare_google_query_addresses.py ciktisini "Bina No Var mi" kolonuna gore
ikiye ayirir: gecoding'e gonderilecek satirlar ve bina numarasi eksik oldugu
icin once adres duzeltmesi gereken satirlar.

Kullanim:
    python scripts/split_missing_house_number.py --input "<GoogleQueryHazir.xlsx>" --ready-output "<hazir.xlsx>" --missing-output "<eksik.xlsx>"
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--ready-output", required=True, type=Path)
    parser.add_argument("--missing-output", required=True, type=Path)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    has_no_col = idx["Bina No Var mi"]

    rows = list(rows_iter)
    ready = [r for r in rows if r[has_no_col] == "Evet"]
    missing = [r for r in rows if r[has_no_col] != "Evet"]

    def write(path: Path, data: list[tuple]) -> None:
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.append(header)
        for row in data:
            out_ws.append(list(row))
        out_wb.save(path)

    write(args.ready_output, ready)
    write(args.missing_output, missing)

    print(f"Toplam: {len(rows)}")
    print(f"Bina no ile hazir (geocode edilecek): {len(ready)} -> {args.ready_output}")
    print(f"Bina no eksik (adres duzeltmesi gerekiyor): {len(missing)} -> {args.missing_output}")


if __name__ == "__main__":
    main()
