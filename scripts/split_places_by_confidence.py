#!/usr/bin/env python3
"""
places_retry_ekkonum.py ciktilarini uc kategoriye ayirir:
  - temiz: places_status=OK VE beklenen ilce, formatted_address icinde geciyor
  - supheli: OK ama ilce uyusmuyor
  - sonuc_yok: ZERO_RESULTS / bos sorgu / diger hatalar

Kullanim:
    python scripts/split_places_by_confidence.py --input "<...Sonuc.xlsx>" --clean-output "<temiz.xlsx>" --suspicious-output "<supheli.xlsx>" --empty-output "<sonucsuz.xlsx>"
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

TURKISH_MAP = str.maketrans("İÇĞÖŞÜ", "ICGOSU")


def norm(value: str) -> str:
    return value.upper().translate(TURKISH_MAP)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--clean-output", required=True, type=Path)
    parser.add_argument("--suspicious-output", required=True, type=Path)
    parser.add_argument("--empty-output", required=True, type=Path)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    rows = list(rows_iter)

    clean, suspicious, empty = [], [], []
    for row in rows:
        status = row[idx["places_status"]]
        if status != "OK":
            empty.append(row)
            continue

        address = str(row[idx["Adres"]] or "")
        segments = [s.strip() for s in address.split(",") if s.strip()]
        ilce = segments[-2] if len(segments) >= 2 else ""
        formatted = str(row[idx["places_formatted_address"]] or "")

        if ilce and norm(ilce) not in norm(formatted):
            suspicious.append(row)
        else:
            clean.append(row)

    def write(path: Path, data: list[tuple]) -> None:
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.append(header)
        for row in data:
            out_ws.append(list(row))
        out_wb.save(path)

    write(args.clean_output, clean)
    write(args.suspicious_output, suspicious)
    write(args.empty_output, empty)

    print(f"Toplam: {len(rows)}")
    print(f"Temiz: {len(clean)} -> {args.clean_output}")
    print(f"Şüpheli: {len(suspicious)} -> {args.suspicious_output}")
    print(f"Sonuçsuz: {len(empty)} -> {args.empty_output}")


if __name__ == "__main__":
    main()
