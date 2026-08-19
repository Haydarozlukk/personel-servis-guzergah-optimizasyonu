#!/usr/bin/env python3
"""
AdresEksik.xlsx'i ciddiyete gore ikiye ayirir:
  - orta: Mahalle/Cadde-Sokak bilgisi var, sadece Bina No eksik
  - kritik: Google Sorgu Adresi'nde sadece Ilce/Il kaldi (form neredeyse bos)

Kullanim:
    python scripts/split_missing_by_severity.py --input "<AdresEksik.xlsx>" --moderate-output "<orta.xlsx>" --critical-output "<kritik.xlsx>"
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--moderate-output", required=True, type=Path)
    parser.add_argument("--critical-output", required=True, type=Path)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    rows = list(rows_iter)

    moderate: list[tuple] = []
    critical: list[tuple] = []
    for row in rows:
        query = str(row[idx["Google Sorgu Adresi"]] or "").strip()
        segments = [s for s in (part.strip() for part in query.split(",")) if s]
        (critical if len(segments) <= 2 else moderate).append(row)

    def write(path: Path, data: list[tuple]) -> None:
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.append(header)
        for row in data:
            out_ws.append(list(row))
        out_wb.save(path)

    write(args.moderate_output, moderate)
    write(args.critical_output, critical)

    print(f"Toplam: {len(rows)}")
    print(f"Orta (sadece bina no eksik): {len(moderate)} -> {args.moderate_output}")
    print(f"Kritik (adres neredeyse bos): {len(critical)} -> {args.critical_output}")


if __name__ == "__main__":
    main()
