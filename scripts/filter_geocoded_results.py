#!/usr/bin/env python3
"""
Geocoded_v3.xlsx ciktisini bina no + location_type kalitesine gore uce ayirir:
  - ready: Bina No var VE location_type guvenilir (ROOFTOP/RANGE_INTERPOLATED/GEOMETRIC_CENTER)
  - review: location_type APPROXIMATE (Bina No olsa da olmasa da supheli)
  - missing: Bina No hic yoktu (zaten adres eksikti)

Kullanim:
    python scripts/filter_geocoded_results.py --input "<Geocoded_v3.xlsx>" --ready-output "<hazir.xlsx>" --review-output "<gozden_gecir.xlsx>" --missing-output "<eksik.xlsx>"
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

RELIABLE_TYPES = {"ROOFTOP", "RANGE_INTERPOLATED", "GEOMETRIC_CENTER"}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--ready-output", required=True, type=Path)
    parser.add_argument("--review-output", required=True, type=Path)
    parser.add_argument("--missing-output", required=True, type=Path)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    has_no_col = idx["Bina No Var mi"]
    lt_col = idx["location_type"]

    rows = list(rows_iter)
    missing = [r for r in rows if r[has_no_col] != "Evet"]
    with_no = [r for r in rows if r[has_no_col] == "Evet"]
    review = [r for r in with_no if r[lt_col] not in RELIABLE_TYPES]
    ready = [r for r in with_no if r[lt_col] in RELIABLE_TYPES]

    def write(path: Path, data: list[tuple]) -> None:
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.append(header)
        for row in data:
            out_ws.append(list(row))
        out_wb.save(path)

    write(args.ready_output, ready)
    write(args.review_output, review)
    write(args.missing_output, missing)

    print(f"Toplam: {len(rows)}")
    print(f"Hazir (bina no + guvenilir konum): {len(ready)} -> {args.ready_output}")
    print(f"Gozden gecirilecek (APPROXIMATE): {len(review)} -> {args.review_output}")
    print(f"Adres eksik (bina no yok): {len(missing)} -> {args.missing_output}")


if __name__ == "__main__":
    main()
