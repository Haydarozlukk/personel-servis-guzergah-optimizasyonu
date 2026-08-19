#!/usr/bin/env python3
"""
Kritik_Supheli, Kritik_Sonucsuz, Orta_Supheli, Orta_Sonucsuz dosyalarini ve
Istanbul'da cikan 2 kisiyi tek bir 'uzmana git' Excel'inde birlestirir, her
satira neden flag'lendigini aciklayan bir 'Sebep' kolonu ekler.

Kullanim:
    python scripts/build_expert_review_list.py --output "<UzmanaGit.xlsx>"
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path

import openpyxl

DOWNLOADS = Path(os.environ.get("USERPROFILE", "")) / "Downloads"

SOURCES = [
    ("Kritik_Supheli.xlsx", "Şüpheli: Google başka ilçeye/bölgeye eşleşti"),
    ("Kritik_Sonucsuz.xlsx", "Sonuçsuz: hiçbir bilgi yok ya da Google bulamadı"),
    ("Orta_Supheli.xlsx", "Şüpheli: Google başka ilçeye/bölgeye eşleşti"),
    ("Orta_Sonucsuz.xlsx", "Sonuçsuz: hiçbir bilgi yok ya da Google bulamadı"),
]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    out_header = ["Ad Soyad", "Sicil", "Birim", "Adres", "Kullandığı Servis", "Sebep"]
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(out_header)

    total = 0
    for filename, reason in SOURCES:
        path = DOWNLOADS / filename
        if not path.exists():
            print(f"Atlandı (bulunamadı): {path}")
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
        idx = {h: i for i, h in enumerate(header)}
        for row in rows_iter:
            out_ws.append([
                row[idx.get("Ad Soyad", -1)] if "Ad Soyad" in idx else "",
                row[idx.get("Sicil", -1)] if "Sicil" in idx else "",
                row[idx.get("Birim", -1)] if "Birim" in idx else "",
                row[idx.get("Adres", -1)] if "Adres" in idx else "",
                row[idx.get("Kullandığı Servis", -1)] if "Kullandığı Servis" in idx else "",
                reason,
            ])
            total += 1

    # Istanbul'da cikan 2 kisi: kayit hatasi degil, ikisinin de "Kullandigi
    # Servis" alaninda zaten ayri (Istanbul) bir servis adi kayitli --
    # bu Ankara planina hic girmemeliler, ayri bir sisteme ait.
    out_ws.append([
        "Ali UZEL", "1060", "4 NOLU UYG DAİ BAŞK.",
        "ERTUĞRULGAZİ Caddesi, İç Kapı (Daire) No: 11, Ek Konum: KALAMIŞ Sitesi D 42 Blok, Başakşehir, İstanbul",
        "BAŞAKŞEHİR",
        "Kayıt hatası değil: Kullandığı Servis zaten 'BAŞAKŞEHİR' — İstanbul'daki ayrı bir servis sistemine ait, bu Ankara planına dahil edilmemeli.",
    ])
    out_ws.append([
        "ELİF ATAGÜNDÜZ", "627", "TOPLU KONUT PROJELERİ VE ARAŞTIRMA DAİRESİ BAŞKANLIĞI",
        "AZERBAYCAN Caddesi, (3. Caddesi, ) YENİ 39. Sokak, Bina No: 14, Bahçelievler, İstanbul",
        "BAHÇELİEVLER",
        "Kayıt hatası değil: Kullandığı Servis zaten 'BAHÇELİEVLER' — İstanbul'daki ayrı bir servis sistemine ait, bu Ankara planına dahil edilmemeli.",
    ])
    total += 2

    out_wb.save(args.output)
    print(f"Toplam: {total}")
    print(f"Çıktı: {args.output}")


if __name__ == "__main__":
    main()
