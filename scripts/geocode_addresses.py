#!/usr/bin/env python3
"""
Google Geocoding API ile toplu adres -> koordinat donusumu (tek seferlik toplu is).

Kullanim:
    set GOOGLE_MAPS_API_KEY=AIza...                 (PowerShell: $env:GOOGLE_MAPS_API_KEY="AIza...")
    python scripts/geocode_addresses.py --input samples/Servis_Adresleri_Geocoder_Dogrulanmis.xlsx --output cikti.xlsx

Notlar:
    - API anahtari kod icine YAZILMAZ; ortam degiskeninden ya da --api-key ile okunur.
    - Ayni adres birden fazla satirda geciyorsa tek kez sorgulanir (cache), gereksiz maliyet olusmaz.
    - Rate limit icin istekler arasinda --sleep saniye beklenir (varsayilan Google'in
      varsayilan kotasinin cok altinda, guvenli taraf).
    - Sonuc dosyasinda orijinal kolonlara ek olarak lat, lon, google_formatted_address,
      geocode_status kolonlari eklenir. Bulunamayan/hatali satirlar status kolonundan
      Excel'de filtrelenip elle kontrol edilebilir.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"
COST_PER_REQUEST_USD = 0.005  # Google Geocoding API liste fiyati (2026 basi)
MAX_RETRIES = 3


def geocode_one(address: str, api_key: str, region: str) -> dict:
    params = {
        "address": address,
        "key": api_key,
        "region": region,
        "components": "country:TR",
        "language": "tr",
    }
    url = f"{GEOCODE_URL}?{urllib.parse.urlencode(params)}"

    last_error = ""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            with urllib.request.urlopen(url, timeout=10) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except (urllib.error.URLError, TimeoutError) as exc:
            last_error = str(exc)
            time.sleep(1.5 * attempt)
            continue

        status = payload.get("status", "UNKNOWN_ERROR")
        if status == "OK":
            result = payload["results"][0]
            location = result["geometry"]["location"]
            return {
                "lat": location["lat"],
                "lon": location["lng"],
                "formatted_address": result.get("formatted_address", ""),
                "location_type": result["geometry"].get("location_type", ""),
                "status": "OK",
            }
        if status == "ZERO_RESULTS":
            return {"lat": None, "lon": None, "formatted_address": "", "location_type": "", "status": "ZERO_RESULTS"}
        if status in ("OVER_QUERY_LIMIT", "UNKNOWN_ERROR"):
            last_error = status
            time.sleep(2.0 * attempt)
            continue
        # REQUEST_DENIED, INVALID_REQUEST vb. tekrar denenmesi anlamsiz hatalar
        return {"lat": None, "lon": None, "formatted_address": "", "location_type": "", "status": status}

    return {"lat": None, "lon": None, "formatted_address": "", "location_type": "", "status": f"FAILED: {last_error}"}


def read_rows(path: Path) -> tuple[list[str], list[dict]]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
        rows = [dict(zip(header, row)) for row in rows_iter]
        return header, rows
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            return list(reader.fieldnames or []), list(reader)
    raise ValueError(f"Desteklenmeyen dosya turu: {path.suffix}")


def write_rows(path: Path, header: list[str], rows: list[dict]) -> None:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(header)
        for row in rows:
            ws.append([row.get(col, "") for col in header])
        wb.save(path)
        return
    if path.suffix.lower() == ".csv":
        with path.open("w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=header)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        return
    raise ValueError(f"Desteklenmeyen dosya turu: {path.suffix}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path, help="Girdi .xlsx/.csv dosyasi")
    parser.add_argument("--output", required=True, type=Path, help="Cikti .xlsx/.csv dosyasi")
    parser.add_argument("--address-column", default="adres", help="Adres kolonu adi (varsayilan: adres)")
    parser.add_argument("--api-key", default=os.environ.get("GOOGLE_MAPS_API_KEY", ""))
    parser.add_argument("--region", default="tr")
    parser.add_argument("--sleep", type=float, default=0.1, help="Istekler arasi bekleme (sn)")
    args = parser.parse_args()

    if not args.api_key:
        sys.exit("API anahtari yok. GOOGLE_MAPS_API_KEY ortam degiskenini ayarlayin ya da --api-key verin.")

    header, rows = read_rows(args.input)
    if args.address_column not in header:
        sys.exit(f"'{args.address_column}' kolonu bulunamadi. Mevcut kolonlar: {header}")

    out_header = [*header, "lat", "lon", "google_formatted_address", "location_type", "geocode_status"]
    cache: dict[str, dict] = {}
    request_count = 0
    not_found = []

    for i, row in enumerate(rows, start=1):
        address = str(row.get(args.address_column, "") or "").strip()
        if not address:
            row.update({"lat": None, "lon": None, "google_formatted_address": "", "geocode_status": "EMPTY_ADDRESS"})
            continue

        if address in cache:
            result = cache[address]
        else:
            result = geocode_one(address, args.api_key, args.region)
            cache[address] = result
            request_count += 1
            time.sleep(args.sleep)

        row["lat"] = result["lat"]
        row["lon"] = result["lon"]
        row["google_formatted_address"] = result["formatted_address"]
        row["location_type"] = result["location_type"]
        row["geocode_status"] = result["status"]

        if result["status"] != "OK":
            not_found.append((i, address, result["status"]))

        print(f"[{i}/{len(rows)}] {address[:60]!r} -> {result['status']}")

    write_rows(args.output, out_header, rows)

    rooftop_count = sum(1 for row in rows if row.get("location_type") == "ROOFTOP")

    print(f"\nToplam satir: {len(rows)}")
    print(f"Google'a atilan benzersiz sorgu: {request_count} (~${request_count * COST_PER_REQUEST_USD:.2f})")
    print(f"ROOFTOP (tam bina eslesmesi): {rooftop_count}/{len(rows)}")
    print(f"Bulunamayan/hatali: {len(not_found)}")
    for i, address, status in not_found:
        print(f"  satir {i}: {status} -> {address}")
    print(f"\nCikti yazildi: {args.output}")


if __name__ == "__main__":
    main()
