#!/usr/bin/env python3
"""
Adresteki "Ek Konum: ..." (site/apartman/blok adı gibi POI bilgisi) segmentini
Google sorgu adresinden ayırır. Bu segment Google'a gonderilen metne karisinca
(orn. "Karaman Suit") aramayi bina numarasindan bagimsiz bir yer adina
kaydirabiliyor; bu yuzden sorguda yer almamali, ama insan icin referans olarak
ayri bir kolonda saklanir.

Kullanim:
    python scripts/prepare_google_query_addresses.py --input "<girdi.xlsx>" --output "<cikti.xlsx>" [--sheet "Servis Bilgileri"]
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl

# "Ek Konum:", "İç Kapı (Daire) No:", "Daire No:", "Kapı No:" gibi segmentler
# Google sorgusunu bozuyor: apartman/site adlari sorguyu alakasiz bir POI'ye
# kaydirabiliyor, ic kapi/daire numaralari da bina numarasiyla karisip
# Nominatim/Google'in numarayi hic tanimamasina yol acabiliyor. Sorgudan
# tamamen cikarilir, referans icin ayri kolonda saklanir.
DROPPED_SEGMENT_PATTERN = re.compile(
    r"^(?:ek konum|i[cç] kap[ıi]\s*\(daire\)\s*no|daire no|kap[ıi] no)\s*:",
    re.IGNORECASE,
)


def split_google_query(address: str) -> tuple[str, str]:
    """Adresi (google_sorgu_adresi, cikarilan_parcalar) ikilisine ayirir."""
    if not address:
        return "", ""

    segments = [segment.strip() for segment in address.split(",")]
    kept: list[str] = []
    dropped_parts: list[str] = []

    for segment in segments:
        if DROPPED_SEGMENT_PATTERN.match(segment):
            dropped_parts.append(DROPPED_SEGMENT_PATTERN.sub("", segment).strip())
        else:
            kept.append(segment)

    google_query = ", ".join(part for part in kept if part)
    dropped = ", ".join(part for part in dropped_parts if part)
    return google_query, dropped


def has_house_number(google_query: str) -> bool:
    return bool(re.search(r"\bbina no\s*:?\s*\d", google_query, re.IGNORECASE))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--sheet", default="Servis Bilgileri")
    parser.add_argument("--address-column", default="Adres")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb[args.sheet]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    if args.address_column not in header:
        raise SystemExit(f"'{args.address_column}' kolonu bulunamadi. Mevcut kolonlar: {header}")

    rows = [dict(zip(header, row)) for row in rows_iter]

    out_header = [*header, "Google Sorgu Adresi", "Çıkarılan Bilgi (Ek Konum/Daire No)", "Bina No Var mi"]
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(out_header)

    no_house_number = []
    dropped_count = 0

    for i, row in enumerate(rows, start=1):
        address = str(row.get(args.address_column, "") or "").strip()
        google_query, dropped = split_google_query(address)
        has_no = has_house_number(google_query)

        if dropped:
            dropped_count += 1
        if not has_no:
            no_house_number.append((i, row.get("Ad Soyad", ""), address))

        out_ws.append([*[row.get(col, "") for col in header], google_query, dropped, "Evet" if has_no else "Hayır"])

    out_wb.save(args.output)

    print(f"Toplam satir: {len(rows)}")
    print(f"'Ek Konum/Daire No' ayrilan satir: {dropped_count}")
    print(f"Bina numarasi olmayan (dikkat gerektiren) satir: {len(no_house_number)}")
    for i, name, address in no_house_number:
        print(f"  satir {i}: {name!r} -> {address!r}")
    print(f"\nCikti yazildi: {args.output}")


if __name__ == "__main__":
    main()
