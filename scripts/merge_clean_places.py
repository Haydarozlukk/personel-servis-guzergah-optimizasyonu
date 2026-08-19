#!/usr/bin/env python3
"""
Iki 'temiz' Places sonucunu (Kritik_Temiz.xlsx, Orta_Temiz.xlsx) ve mevcut
ImportReady_v2.xlsx'i tek bir 'sicil numarasi, ad soyad, adres, boylam, enlem'
dosyasinda birlestirir.

Kullanim:
    python scripts/merge_clean_places.py --existing "<ImportReady_v2.xlsx>" --extra "<Kritik_Temiz.xlsx>" "<Orta_Temiz.xlsx>" --output "<Birlesik.xlsx>"
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def read_existing(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # header
    return [list(row) for row in rows_iter]


def read_extra(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    result = []
    for row in rows_iter:
        sicil = row[idx["Sicil"]]
        name = row[idx["Ad Soyad"]]
        adres = row[idx["Adres"]]
        lon = row[idx["places_lon"]]
        lat = row[idx["places_lat"]]
        if sicil and lon is not None and lat is not None:
            result.append([sicil, name, adres, lon, lat])
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--existing", required=True, type=Path)
    parser.add_argument("--extra", required=True, nargs="+", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    existing = read_existing(args.existing)
    extra: list[list] = []
    for path in args.extra:
        extra.extend(read_extra(path))

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "personel"
    ws.append(["sicil numarası", "ad soyad", "adres", "boylam", "enlem"])
    for row in existing:
        ws.append(row)
    for row in extra:
        ws.append(row)
    out_wb.save(args.output)

    print(f"Mevcut: {len(existing)}")
    print(f"Eklenen (Places temiz): {len(extra)}")
    print(f"Toplam: {len(existing) + len(extra)}")
    print(f"Çıktı: {args.output}")


if __name__ == "__main__":
    main()
