#!/usr/bin/env python3
"""
Geocoding API'nin ayni koordinata dusurdugu (mahalle/kompleks merkezine
"cakisan") kayitlar icin Places API - Text Search ile yeniden sorgu.

Text Search, Maps arama kutusuyla ayni motoru kullanir; "TOKI Konutlari X
Blok" gibi isimlendirilmis site/blok adlarini POI olarak tanima ihtimali
Geocoding API'den yuksektir.

Kullanim:
    python scripts/places_retry.py --input "<geocoded_xlsx>" --output "<retry_output.xlsx>"

Girdi, geocode_addresses.py'nin urettigi dosya olmali (lat/lon/geocode_status
kolonlari icerir). Script once ayni koordinata dusen (>=2 kisi) satirlari
bulur, sadece onlari Places Text Search ile yeniden sorgular.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

PLACES_URL = "https://maps.googleapis.com/maps/api/place/textsearch/json"
COST_PER_REQUEST_USD = 0.032
MAX_RETRIES = 3


def places_search_one(query: str, api_key: str, region: str) -> dict:
    params = {"query": query, "key": api_key, "region": region, "language": "tr"}
    url = f"{PLACES_URL}?{urllib.parse.urlencode(params)}"

    last_error = ""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            with urllib.request.urlopen(url, timeout=10) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except (urllib.error.URLError, TimeoutError) as exc:
            last_error = str(exc)
            time.sleep(1.5 * attempt)
            continue

        status = payload.get("status", "UNKNOWN_ERROR")
        if status == "OK":
            result = payload["results"][0]
            location = result["geometry"]["location"]
            return {
                "lat": location["lat"],
                "lon": location["lng"],
                "name": result.get("name", ""),
                "formatted_address": result.get("formatted_address", ""),
                "status": "OK",
            }
        if status == "ZERO_RESULTS":
            return {"lat": None, "lon": None, "name": "", "formatted_address": "", "status": "ZERO_RESULTS"}
        if status in ("OVER_QUERY_LIMIT", "UNKNOWN_ERROR"):
            last_error = status
            time.sleep(2.0 * attempt)
            continue
        return {"lat": None, "lon": None, "name": "", "formatted_address": "", "status": status}

    return {"lat": None, "lon": None, "name": "", "formatted_address": "", "status": f"FAILED: {last_error}"}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--api-key", default=os.environ.get("GOOGLE_MAPS_API_KEY", ""))
    parser.add_argument("--region", default="tr")
    parser.add_argument("--sleep", type=float, default=0.15)
    args = parser.parse_args()

    if not args.api_key:
        sys.exit("API anahtari yok. GOOGLE_MAPS_API_KEY ortam degiskenini ayarlayin ya da --api-key verin.")

    import openpyxl

    wb = openpyxl.load_workbook(args.input)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    rows = [dict(zip(header, row)) for row in rows_iter]

    # Ayni koordinata (6 hane yuvarlanmis) dusen 2+ kisilik gruplari bul.
    groups: dict[tuple[float, float], list[int]] = defaultdict(list)
    for i, row in enumerate(rows):
        lat, lon = row.get("lat"), row.get("lon")
        if lat is None or lon is None:
            continue
        key = (round(float(lat), 6), round(float(lon), 6))
        groups[key].append(i)

    affected_indexes = {i for indexes in groups.values() if len(indexes) > 1 for i in indexes}
    print(f"Ayni koordinatta 2+ kisi bulunan toplam satir: {len(affected_indexes)}")

    out_header = [*header, "places_lat", "places_lon", "places_name", "places_formatted_address", "places_status"]
    request_count = 0
    moved_count = 0

    for i, row in enumerate(rows):
        if i not in affected_indexes:
            row["places_lat"] = row["places_lon"] = row["places_name"] = row["places_formatted_address"] = row["places_status"] = ""
            continue

        query = str(row.get("Adres", "") or "").strip()
        if not query:
            row["places_status"] = "EMPTY_ADDRESS"
            continue

        result = places_search_one(query, args.api_key, args.region)
        request_count += 1
        time.sleep(args.sleep)

        row["places_lat"] = result["lat"]
        row["places_lon"] = result["lon"]
        row["places_name"] = result["name"]
        row["places_formatted_address"] = result["formatted_address"]
        row["places_status"] = result["status"]

        old_key = (round(float(row["lat"]), 6), round(float(row["lon"]), 6))
        if result["lat"] is not None:
            new_key = (round(result["lat"], 6), round(result["lon"], 6))
            if new_key != old_key:
                moved_count += 1

        print(f"[{request_count}/{len(affected_indexes)}] {row.get('Ad Soyad','')!r} -> {result['status']} "
              f"{'(farkli nokta)' if result['lat'] and (round(result['lat'],6), round(result['lon'],6)) != old_key else ''}")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(out_header)
    for row in rows:
        out_ws.append([row.get(col, "") for col in out_header])
    out_wb.save(args.output)

    print(f"\nSorgulanan satir: {request_count} (~${request_count * COST_PER_REQUEST_USD:.2f})")
    print(f"Eskisinden farkli bir noktaya tasinan: {moved_count}")
    print(f"Cikti: {args.output}")


if __name__ == "__main__":
    main()
